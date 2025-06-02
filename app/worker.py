import pika
import requests
import base64
import rarfile
import openpyxl
import io
import logging
from config import LOG_FILE, API_URL, RABBITMQ_URL
from database import update_report_status, get_jwt_token, get_active_organizations, save_realization_weekly

# Настройка логгера
logging.basicConfig(
    filename=LOG_FILE,
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s"
)
logger = logging.getLogger(__name__)


def process_realization_excel(ch, method, properties, body):
    report_id = body.decode()
    logger.info(f"Получен report_id для обработки: {report_id}")

    try:
        # Пример: получаем токен для первой активной организации
        org_ids = get_active_organizations()
        if not org_ids:
            logger.error("Нет активных организаций")
            ch.basic_ack(delivery_tag=method.delivery_tag)
            return

        jwt = get_jwt_token(org_ids[0])
        if not jwt:
            logger.error(f"Не найден JWT для организации {org_ids[0]}")
            ch.basic_ack(delivery_tag=method.delivery_tag)
            return

        headers = {"Authorization": f"Bearer {jwt}"}
        download_url = f"{API_URL}/download/{report_id}"
        response = requests.get(download_url, headers=headers)

        if response.status_code != 200:
            logger.error(f"Ошибка загрузки отчёта {report_id}, код: {response.status_code}")
            ch.basic_nack(delivery_tag=method.delivery_tag, requeue=False)
            return

        data = response.json()
        b64_content = data.get("data")

        if not b64_content:
            logger.error(f"Пустой или неверный контент в отчёте {report_id}")
            ch.basic_nack(delivery_tag=method.delivery_tag, requeue=False)
            return

        # Декодируем Base64
        rar_data = base64.b64decode(b64_content)

        # Распаковываем RAR
        with rarfile.RarFile(io.BytesIO(rar_data)) as rf:
            if '0.xlsx' not in [f.filename for f in rf.infolist()]:
                logger.error(f"Файл 0.xlsx не найден в архиве отчёта {report_id}")
                ch.basic_nack(delivery_tag=method.delivery_tag, requeue=False)
                return

            xlsx_data = rf.read('0.xlsx')

        # Парсим Excel
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_data))
        ws = wb.active

        # Здесь можно добавить запись в БД
        for row in ws.iter_rows(min_row=2):  # пропускаем заголовок
            values = [cell.value for cell in row]
            logger.debug(f"Обработаны данные: {values}")
            wb = openpyxl.load_workbook(io.BytesIO(xlsx_data))
            ws = wb.active

            rows = []
            for row in ws.iter_rows(min_row=2):  # пропускаем заголовок
                values = [cell.value for cell in row]
                rows.append(tuple(values))

            # Сохраняем в БД
            save_realization_weekly(rows)
            logger.info(f"{len(rows)} строк успешно записано в wildberries.reports.realization_weekly")

        # Отправляем в следующую очередь
        connection = pika.BlockingConnection(pika.URLParameters(RABBITMQ_URL))
        channel = connection.channel()
        channel.queue_declare(queue='xls_to_excel', durable=True)

        message = f"{report_id}|wildberries.reports.realization_weekly"
        channel.basic_publish(
            exchange='files',
            routing_key='realization_excel',
            body=message,
            properties=pika.BasicProperties(delivery_mode=2)  # сообщение устойчивое
        )
        connection.close()

        # Обновляем статус в БД
        update_report_status(report_id, status="downloaded")
        logger.info(f"Отчёт {report_id} успешно обработан и отправлен в очередь files.xls_to_excel")

        ch.basic_ack(delivery_tag=method.delivery_tag)

    except Exception as e:
        logger.error(f"Ошибка при обработке отчёта {report_id}: {str(e)}", exc_info=True)
        ch.basic_nack(delivery_tag=method.delivery_tag, requeue=False)


def start_worker():
    logger.info("Запуск RabbitMQ-воркера для очереди realization_excel")

    connection = pika.BlockingConnection(pika.URLParameters(RABBITMQ_URL))
    channel = connection.channel()

    channel.queue_declare(queue='realization_excel', durable=True)

    channel.basic_consume(
        queue='realization_excel',
        on_message_callback=process_realization_excel,
        auto_ack=False
    )

    logger.info("Воркер запущен. Ожидание сообщений...")
    try:
        channel.start_consuming()
    except KeyboardInterrupt:
        logger.info("Воркер остановлен вручную.")
        connection.close()

if __name__ == "__main__":
    start_worker()